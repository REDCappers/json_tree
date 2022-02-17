import json

import openpyxl
from openpyxl.styles import Side, Border, PatternFill

# Excel
wb = openpyxl.Workbook()
sheet = wb.active
sheet.title = 'result'

# 罫線
solid = Side(border_style="thin", color="000000")

# 縦線の線種と位置の定義（縦線は列の左だけ線を引く）
border_start = Border(top=solid, bottom=solid, right=None, left=solid)
border_middle = Border(top=solid, bottom=solid, left=None, right=None)
border_end = Border(top=solid, bottom=solid, left=None, right=solid)

gyou = 1

# テスト
start_fill = PatternFill(patternType='solid', fgColor='AAFFFF')
end_fill = PatternFill(patternType='solid', fgColor='FFAAFF')


# セルに罫線を設定（value用）
def border_draw_range(row, start, end):
    for col_num in range(start, end + 1):
        if col_num == start:
            sheet.cell(row=row, column=col_num).border = border_start
        elif col_num == end:
            sheet.cell(row=row, column=col_num).border = border_end
        else:
            sheet.cell(row=row, column=col_num).border = border_middle


# セルに罫線を設定（list用）
def border_draw_list_range(row, start, end):
    for col_num in range(start, end + 1):
        if col_num == start:
            sheet.cell(row=row, column=col_num).border = Border(left=None, bottom=solid, top=solid)
        elif col_num == end:
            sheet.cell(row=row, column=col_num).border = Border(left=None, bottom=solid, top=solid, right=solid)
        else:
            sheet.cell(row=row, column=col_num).border = border_middle


# セルに罫線を設定（list内部用）
def border_draw_list_end(row, start, end):
    """
    終端用の色を塗る. 最初、末尾、間で処理分けているが、特に処理は変えていない.
    :param row:
    :param start:
    :param end:
    :return:
    """
    for col_num in range(start, end + 1):
        if col_num == start:
            sheet.cell(row=row, column=col_num).fill = end_fill
        elif col_num == end:
            sheet.cell(row=row, column=col_num).fill = end_fill
        else:
            sheet.cell(row=row, column=col_num).fill = end_fill


# セルに罫線を設定（list内部用）
def border_draw_list_start(row, start, end):
    """
    開始用の色を塗る. 最初、末尾、間で処理分けているが、特に処理は変えていない.
    :param row:
    :param start:
    :param end:
    :return:
    """
    for col_num in range(start, end + 1):
        if col_num == start:
            sheet.cell(row=row, column=col_num).fill = start_fill
        elif col_num == end:
            sheet.cell(row=row, column=col_num).fill = start_fill
        else:
            sheet.cell(row=row, column=col_num).fill = start_fill


def json_tree(data, retsu=1):
    """
    JSONを階層構造にintendを行う
    :param data:
    :param retsu:
    :return:
    """
    global sheet
    global gyou
    if type(data) == dict:
        for k in data.keys():
            gyou += 1
            sheet.cell(row=gyou, column=retsu, value=k)
            if type(data[k]) == dict:
                gyou -= 1

            json_tree(data[k], retsu + 1)

    elif type(data) == list and len(data) > 0:
        sheet.cell(row=gyou, column=retsu)
        for d in data:
            json_tree(d, retsu)  # 改行なし
    else:
        sheet.cell(row=gyou, column=retsu, value=data)
    return


def draw_vertical_line(max_row, max_column):
    """
    全体の結果を枠線で囲む
    :param max_row:
    :param max_column:
    :return:
    """
    global sheet
    for row_num in range(1, max_row + 1):
        for col_num in range(1, max_column + 1):
            sheet.cell(column=col_num, row=row_num).border = Border(left=solid, right=solid)
            if row_num == max_row:
                sheet.cell(column=col_num, row=row_num).border = Border(left=solid, right=solid, bottom=solid)


def draw_outline(max_row, max_column):
    """
    外枠を描画. 中の枠線が上書きされないよう大外のセルを指定して描画.
    :param max_row:
    :param max_column:
    :return:
    """
    global sheet
    for row_num in range(1, max_row + 2):
        for col_num in range(1, max_column + 2):
            if row_num == max_row + 1:
                sheet.cell(column=col_num, row=row_num).border = Border(top=solid)
            if col_num == max_column + 1:
                sheet.cell(column=col_num, row=row_num).border = Border(left=solid)
            if col_num == max_column + 1 and row_num == max_row + 1:
                sheet.cell(column=col_num, row=row_num).border = Border(left=None, top=None)


def draw_tree(data, retsu=1):
    """
    階層構造の枠線を作成.
    :param data:
    :param retsu:
    :return:
    """
    global sheet
    global gyou
    if type(data) == dict:
        for k in data.keys():
            gyou += 1
            sheet.cell(row=gyou, column=retsu).border = Border(top=solid, left=solid)
            if type(data[k]) == dict:
                gyou -= 1

            # 配列の先頭、末尾判定
            if next(iter(reversed(data))) == k:
                border_draw_list_end(gyou, retsu, sheet.max_column)
            elif next(iter(data)) == k:
                border_draw_list_start(gyou, retsu, sheet.max_column)

            draw_tree(data[k], retsu + 1)
    elif type(data) == list and len(data) > 0:
        border_draw_list_range(gyou, retsu, sheet.max_column)
        for d in data:
            draw_tree(d, retsu)  # 改行なし
    else:
        border_draw_range(gyou, retsu, sheet.max_column)
    return


"""
メイン処理
"""
file_path = \
    "C:\\Users\\REDCap\\Desktop\\Questionnaire-QuestionnaireResponse調査\\jsonデータ\\QuestionnaireResponse\\⑥20220209_多発性骨髄腫(KMF)テンプレート診断.json"
with open(file_path, 'r', encoding='utf-8') as json_file:

    # jsonデータ取り込み
    json_load = json.load(json_file)

    # jsonデータ記載
    json_tree(json_load)

    # 縦線作成
    draw_vertical_line(sheet.max_row, sheet.max_column)

    gyou = 1  # 初期化

    # 階層枠線作成
    draw_tree(json_load)

    # 外枠作成
    draw_outline(sheet.max_row, sheet.max_column)

    # 出力
    wb.save('C:\\Users\\REDCap\\Desktop\\temp\\data.xlsx')
